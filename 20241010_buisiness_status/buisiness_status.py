from typing import List, Union
import requests
from datetime import datetime
import openpyxl, xlrd, csv
import chardet
import json
import sys 
from PyQt5.QtWidgets import QMainWindow, QLineEdit, QPushButton, QLabel, QFileDialog, QApplication
from PyQt5.QtCore import Qt

# 파일 정보
class FileInfo:
    def __init__(self, input_file_path:str, output_dir_path:str, column_num:Union[int, str]='사업자등록 번호'):
        self.input_file_path=input_file_path
        self.output_dir_path=output_dir_path
        self.column_num=column_num

# API 요청 바디
class RequestBody:
    def __init__(self):
        self.body=[]
    def iput_bnos(self, buisiness_no_list:List[str]=[]):
        self.body={'b_no':buisiness_no_list}

# API 요청 정보
class RequestParam:
    def __init__(self, base_url:str, service_key:str, request_body=RequestBody):
        self.base_url=base_url
        self.service_key=service_key
        self.request_body=request_body.body
        
# 기본 설정값
class Globals:
    def __init__(self):
        self.base_url = 'http://api.odcloud.kr/api/nts-businessman/v1/status?serviceKey='
        
# API 요청
def lookup_buisiness_status(request_param:RequestParam):
    url = request_param.base_url + request_param.service_key
    headers = {'Content-Type': 'application/json'}
    body = request_param.request_body
    response = requests.post(url=url, headers=headers, json=body.__dict__.get('body'))
    return json.loads(response.content)

# 사업자번호에서 dash(-) 제거
def replace_dash(bno_list:List[str]):
    return [str(bno).replace('-', '') for bno in bno_list]

# 텍스트가 영문 A - z 로만 이루어졌는지 확인
def is_all_alpha(text):
    result = all('A' <= char <= 'z' for char in text)
    return result

# 영문자를 숫자로 바꾸기
def excel_column_to_number(column_name):
    result = 0
    for char in column_name:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

# input 파일 읽고 사업자 등록 번호 리스팅
def listing_buisiness_nos(file_info):
    buisiness_no_list = []
    if file_info.input_file_path.split('.')[-1] == 'xlsx':
        wb = openpyxl.load_workbook(file_info.input_file_path)
        sheet = wb.active
        try:
            b_no_index = int(file_info.column_num) - 1
        except:
            if is_all_alpha(file_info.column_num):
                b_no_index = excel_column_to_number(file_info.column_num)-1
            else:
                b_no_index = list(next(sheet.iter_rows(values_only=True))).index("사업자등록 번호")
        for row in sheet.iter_rows(values_only=True):
            buisiness_no_list.append(row[b_no_index])
    elif file_info.input_file_path.split('.')[-1] == 'xls':
        wb = xlrd.open_workbook(file_info.input_file_path)
        sheet = wb.sheet_by_index(0)
        try:
            b_no_index = int(file_info.column_num) - 1
        except:
            if is_all_alpha(file_info.column_num):
                b_no_index = excel_column_to_number(file_info.column_num)-1
            else:
                b_no_index = sheet.row_values(0).index("사업자등록 번호")
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            buisiness_no_list.append(row[b_no_index])
    elif file_info.input_file_path.split('.')[-1] == 'csv':
        with open(file_info.input_file_path, 'rb') as f:
            raw_data = f.read()
            encoding = chardet.detect(raw_data)['encoding']

        with open(file_info.input_file_path, newline = '', encoding=encoding) as f:
            reader = csv.reader(f)
            try:
                b_no_index = int(file_info.column_num) - 1
            except:
                if is_all_alpha(file_info.column_num):
                    b_no_index = excel_column_to_number(file_info.column_num)-1
                else:
                    b_no_index = next(reader).index("사업자등록 번호")
            for row in reader:
                buisiness_no_list.append(row[b_no_index])
    return buisiness_no_list

# 사업자 상태 조회
def process(file_info, service_key):
    # global 설정 정보
    glbs = Globals()
    
    # Request Param
    request_body = RequestBody()
    request_param = RequestParam(glbs.base_url, service_key, request_body)

    # 파일 읽고 사업자 등록 번호 리스팅
    buisiness_no_list = listing_buisiness_nos(file_info)

    # 요청 (100개씩)
    result = {'b_no':[], 'status':[]}
    r = len(buisiness_no_list)//100
    for i in range(r+1):
        bnos = replace_dash(buisiness_no_list[i*100:(i+1)*100])
        request_body.iput_bnos(bnos)
        request_param.request_body = request_body
        response = lookup_buisiness_status(request_param).get('data')
        for status_info in response:
            result['b_no'].append(status_info['b_no'])
            status = status_info['b_stt'] if status_info['b_stt'] != '' else '정보 없음'
            result['status'].append(status)

    # 출력 파일 만들기
    with open(file_info.output_dir_path + f'/사업자상태조회_{datetime.now().strftime("%Y-%m-%d")}.csv', 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['b_no_origin', 'b_no', 'status'])
        for b_no_origin, b_no, status in zip(buisiness_no_list, result['b_no'], result['status']):
            writer.writerow([b_no_origin, b_no, status])

# line edit
class CustomLineEdit(QLineEdit):
    def __init__(self, placeholder_text, parent=None):
        super().__init__(parent)
        self.place_holder = placeholder_text
        self.setPlaceholderText(placeholder_text)
    
    # 포커스를 얻었을 때 호출되는 메서드 재정의
    def focusInEvent(self, event):
        super().focusInEvent(event)  # 기본 포커스 동작 수행
        self.setAlignment(Qt.AlignRight)
        self.setPlaceholderText('')  # 포커스 시 Placeholder 삭제
    
    # 포커스를 잃었을 때 다시 Placeholder 추가
    def focusOutEvent(self, event):
        super().focusOutEvent(event)  # 기본 포커스 동작 수행
        if not self.text():  # 입력된 텍스트가 없을 경우에만 placeholder 표시
            self.setAlignment(Qt.AlignLeft)
            self.setPlaceholderText(self.place_holder)

# window
class MyWindow(QMainWindow):
    
    def __init__(self):
        super().__init__()
        # 위젯 기본 설정
        self.setGeometry(300, 300, 390, 250)
        self.setWindowTitle('사업자 상태 조회')
        # Input File Path
        self.line_input_path = CustomLineEdit("조회할 사업자등록번호가 포함된 파일 경로 (xlsx, xls, csv)", self)
        self.line_input_path.setAlignment(Qt.AlignLeft)
        self.line_input_path.setGeometry(10, 10, 325, 25)
        self.btn_input_path = QPushButton('찾기', self)
        self.btn_input_path.setGeometry(335, 8, 50, 25)
        self.btn_input_path.clicked.connect(self.open_file_dialog)
        self.btn_input_path.setFocus()
        # Column Number
        self.line_column_num = CustomLineEdit('사업자번호가 포함된 열 번호를 기입해주세요.', self)
        self.line_column_num.setGeometry(10, 45, 370, 25)
        # Output File Path
        self.line_output_dir = CustomLineEdit('결과 파일을 저장할 경로를 입력해주세요.', self)
        self.line_output_dir.setAlignment(Qt.AlignLeft)
        self.line_output_dir.setGeometry(10, 80, 325, 25)
        self.btn_output_path = QPushButton('찾기', self)
        self.btn_output_path.setGeometry(335, 78, 50, 25)
        self.btn_output_path.clicked.connect(self.open_directory_dialog)
        # API Service Key
        self.line_service_key = CustomLineEdit('API 서비스 키를 입력해주세요.', self)
        self.line_service_key.setGeometry(10, 115, 370, 25)
        # 진행 상황 바 (QProgressBar)
        # self.progress_bar = QProgressBar(self)
        # self.progress_bar.setGeometry(10, 150, 370, 25)
        # self.progress_bar.setValue(0)  # 처음에는 0%로 설정
        # 사업자 상태 조회
        self.btn_process = QPushButton('사업자 상태 조회', self)
        self.btn_process.setGeometry(250, 150, 135, 25)
        self.btn_process.clicked.connect(self.doig_process)
        # 작업 상태
        self.process_status = QLabel('작업 대기', self)
        self.process_status.setGeometry(10, 152, 300, 25)
        # 출처 표기
        self.data_from = QLabel('API 출처 : 국세청', self)
        self.data_from.setGeometry(290, 220, 100, 25)
        self.data_from.setStyleSheet("color : grey;")
        
    
    # 파일 경로 선택 대화상자 (사용자가 파일을 선택할 때)
    def open_file_dialog(self):
        file_name, _ = QFileDialog.getOpenFileName(self, '파일 선택', '', 'All Files (*);; Excel Files (*.xlsx *.xls);; CSV Files (*.csv)')
        if file_name:
            self.line_input_path.setText(file_name)  # 선택한 파일 경로를 QLineEdit에 설정
    
    # 디렉토리 선택 대화상자 (사용자가 디렉토리를 선택할 때)
    def open_directory_dialog(self):
        directory = QFileDialog.getExistingDirectory(self, '디렉토리 선택', '')
        if directory:
            self.line_output_dir.setText(directory)  # 선택한 디렉토리 경로를 QLineEdit에 설정
    
    # 사업자 상태 조회
    def doig_process(self):
        self.process_status.setText('작업중..')
        input_file_path = self.line_input_path.text()
        column_num = self.line_column_num.text()
        output_dir_path = self.line_output_dir.text()
        service_key = self.line_service_key.text()
        if input_file_path == '':
            self.process_status.setText('파일을 선택해주세요.')
            self.process_status.setStyleSheet("color : red;")
        elif output_dir_path == '':
            self.process_status.setText('저장 경로를 선택해주세요.')
            self.process_status.setStyleSheet("color : red;")
        elif service_key == '':
            self.process_status.setText('서비스키를 입력해주세요.')
            self.process_status.setStyleSheet("color : red;")
        else:
            try:
                file_info = FileInfo(input_file_path, output_dir_path, column_num)
                process(file_info, service_key)
                self.process_status.setText('조회 완료')
                self.process_status.setStyleSheet("color : blue;")
                return 1
            except:
                self.process_status.setText('오류 발생')
                return 0

app = QApplication(sys.argv)
mywindow = MyWindow()
mywindow.show()
app.exec_()